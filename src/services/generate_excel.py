#!/usr/bin/env python3
"""
generate_excel.py

Generates an Excel workbook (.xlsx) from namespace / model / connection JSON.

For each dashboard in the namespace a sheet is created containing:
  - A pivot table whose data source is the AtScale XMLA/MDX endpoint
  - One chart per dashboard tile, styled according to the worksheet graphType

The pivot table / OLAP connection is injected as raw XML into the xlsx zip
after openpyxl has written the basic workbook skeleton, because openpyxl does
not expose a first-class OLAP pivot-table API.

Dependencies: openpyxl — installed automatically if not present.
"""

import subprocess
import sys

def _ensure(pkg):
    try:
        __import__(pkg)
    except ImportError:
        for extra in (["--user"], ["--break-system-packages"], []):
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "--quiet", pkg] + extra,
                    stderr=subprocess.DEVNULL,
                )
                return
            except subprocess.CalledProcessError:
                continue
        raise RuntimeError(
            f"Could not install '{pkg}'. "
            f"Run: pip3 install {pkg}  (or: pip3 install --user {pkg})"
        )

_ensure("openpyxl")

import argparse
import json
import os
import re
import shutil
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate Excel from namespace")
    parser.add_argument("--namespace-json-file",  required=True)
    parser.add_argument("--model-json-file",      required=True)
    parser.add_argument("--connection-json-file",  required=True)
    parser.add_argument("--connection-name",       required=True)
    parser.add_argument("--target-file",           required=True)
    args = parser.parse_args()

    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart
        from openpyxl.chart.reference import Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required — install with: pip3 install openpyxl",
              file=sys.stderr)
        sys.exit(1)

    with open(args.namespace_json_file)  as f: namespace       = json.load(f)
    with open(args.model_json_file)      as f: models          = json.load(f)
    with open(args.connection_json_file) as f: connections_data = json.load(f)

    connection_name = args.connection_name
    target_file     = args.target_file

    connection = connections_data.get("connections", {}).get(connection_name)
    if not connection:
        print(f"ERROR: connection '{connection_name}' not found in connection file",
              file=sys.stderr)
        sys.exit(1)

    mdx      = connection.get("mdx", {})
    mdx_url  = mdx.get("url", "").rstrip("/")
    org_id   = mdx.get("organization_id", "default")
    catalog  = mdx.get("catalog_name", "")
    users    = connections_data.get("users", {})
    user_key = mdx.get("user", "admin")
    user_obj = users.get(user_key, {})
    username = user_obj.get("username", user_key)
    password = user_obj.get("password", "")

    # installer:true means AtScale is running on the installer port (10502)
    # rather than the default embedded port.  Only append if no port is
    # already present in the URL.
    if connection.get("installer") and ":" not in mdx_url.split("//", 1)[-1]:
        mdx_url = f"{mdx_url}:10502"

    xmla_url    = f"{mdx_url}/xmla/{org_id}"
    conn_string = (
        f"Provider=MSOLAP.8;"
        f"Data Source={xmla_url};"
        f"Initial Catalog={catalog};"
        f"User ID={username};"
        f"Password={password};"
        f"Persist Security Info=True;"
    )

    worksheets = namespace.get("worksheets", {})
    dashboards = namespace.get("dashboards", {})

    # ------------------------------------------------------------------
    # Build the workbook with openpyxl
    # ------------------------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)

    pivot_meta = []  # collect info for post-processing

    HEADER_COLOR   = "1F3864"
    TITLE_COLOR    = "2E74B5"
    TITLE_BG       = "D6E4F0"
    COL_HEADER_BG  = "2E74B5"
    ALT_ROW_BG     = "EBF3FB"
    BANNER_BG      = "E9EFF8"
    TILES_PER_ROW  = 2
    TILE_W         = 9    # columns
    TILE_H         = 22   # rows

    for dash_name, dashboard in dashboards.items():
        tiles = dashboard.get("tiles", [])
        if not tiles:
            continue

        sheet_title = _safe_name(dash_name)
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.showGridLines = False

        # Banner
        ws.row_dimensions[1].height = 28
        ws["B1"] = dash_name
        ws["B1"].font      = Font(bold=True, size=15, color="FFFFFF")
        ws["B1"].alignment = Alignment(vertical="center")
        for col in range(2, 2 + TILES_PER_ROW * (TILE_W + 1)):
            ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=HEADER_COLOR)
        ws.merge_cells(f"B1:{_col(2 + TILES_PER_ROW * (TILE_W + 1) - 1)}1")

        current_row = 3
        tile_in_row = 0

        for tile in tiles:
            ws_name = tile.get("worksheet")
            if not ws_name or ws_name not in worksheets:
                continue

            ws_def     = worksheets[ws_name]
            model_name = ws_def.get("model", "")
            measures   = ws_def.get("measures", [])
            x_axis     = ws_def.get("xAxis")
            graph_type = ws_def.get("graphType", "bar")
            tile_title = ws_def.get("title", ws_name)

            anchor_col = 2 + tile_in_row * (TILE_W + 1)
            anchor_row = current_row

            # Section title
            title_cell = ws.cell(row=anchor_row, column=anchor_col, value=tile_title)
            title_cell.font      = Font(bold=True, size=11, color=TITLE_COLOR)
            title_cell.fill      = PatternFill("solid", fgColor=TITLE_BG)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[anchor_row].height = 18
            ws.merge_cells(
                start_row=anchor_row, start_column=anchor_col,
                end_row=anchor_row,   end_column=anchor_col + TILE_W - 1,
            )

            # Column headers for the pivot data area
            hdr_row = anchor_row + 1
            headers = ([x_axis] if x_axis else []) + measures
            for i, h in enumerate(headers):
                cell = ws.cell(row=hdr_row, column=anchor_col + i, value=h)
                cell.font      = Font(bold=True, size=9, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
                cell.alignment = Alignment(horizontal="center")
                col_width = max(len(h) + 2, 12)
                ws.column_dimensions[_col(anchor_col + i)].width = col_width

            # Placeholder data rows (filled from MDX at runtime via pivot refresh)
            data_start = hdr_row + 1
            data_end   = data_start + 9
            for r in range(data_start, data_end + 1):
                for i in range(len(headers)):
                    cell = ws.cell(row=r, column=anchor_col + i, value="")
                    if r % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)

            # Chart — skipped for text tiles (pivot table alone suffices)
            if graph_type != "text":
                chart = _build_chart(
                    graph_type, tile_title, ws,
                    hdr_row, anchor_col, len(headers), data_end,
                    has_categories=bool(x_axis),
                )
                chart_anchor = ws.cell(row=data_end + 2, column=anchor_col).coordinate
                ws.add_chart(chart, chart_anchor)

            # Record for OLAP injection
            pivot_meta.append({
                "sheet_title":    sheet_title,
                "model":          model_name,
                "measures":       measures,
                "x_axis":         x_axis,
                "tile_title":     tile_title,
                "hdr_row":        hdr_row,
                "data_start":     data_start,
                "data_end":       data_end,
                "anchor_col":     anchor_col,
                "num_headers":    len(headers),
            })

            tile_in_row += 1
            if tile_in_row >= TILES_PER_ROW:
                tile_in_row  = 0
                current_row += TILE_H

        # Footer: MDX endpoint
        footer_row = ws.max_row + 3
        ws.cell(row=footer_row, column=2, value="AtScale MDX endpoint").font = Font(
            bold=True, size=8, color="808080"
        )
        ws.cell(row=footer_row + 1, column=2, value=xmla_url).font = Font(
            size=8, color="A0A0A0"
        )

    # Hidden sheet: connection details
    _write_connections_sheet(wb, connection_name, conn_string, xmla_url, catalog, username)

    # Save base workbook
    Path(target_file).parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_file)

    # Post-process: inject OLAP connection + pivot-table XML into the xlsx zip
    _inject_olap(target_file, pivot_meta, conn_string, catalog, connection_name, models)

    print(f"Wrote Excel workbook to {target_file}")


# ---------------------------------------------------------------------------
# Chart builder
# ---------------------------------------------------------------------------

def _build_chart(graph_type, title, ws, hdr_row, anchor_col,
                 num_cols, data_end, has_categories):
    from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart
    from openpyxl.chart.reference import Reference

    data_start    = hdr_row + 1
    meas_col_min  = anchor_col + (1 if has_categories else 0)
    meas_col_max  = anchor_col + num_cols - 1

    if graph_type == "line":
        chart = LineChart(); chart.style = 10
    elif graph_type in ("pie", "donut") and num_cols <= 2:
        chart = PieChart(); chart.style = 10
    elif graph_type == "area":
        chart = AreaChart(); chart.style = 10
    else:
        chart = BarChart()
        chart.type     = "col"
        chart.grouping = "clustered"
        chart.style    = 10

    chart.title  = title
    chart.width  = 14
    chart.height = 10

    data_ref = Reference(ws,
        min_col=meas_col_min, max_col=meas_col_max,
        min_row=hdr_row,      max_row=data_end,
    )
    chart.add_data(data_ref, titles_from_data=True)

    if has_categories:
        cats = Reference(ws,
            min_col=anchor_col,
            min_row=data_start, max_row=data_end,
        )
        chart.set_categories(cats)

    return chart


# ---------------------------------------------------------------------------
# Connections reference sheet
# ---------------------------------------------------------------------------

def _write_connections_sheet(wb, connection_name, conn_string,
                              xmla_url, catalog, username):
    from openpyxl.styles import Font, PatternFill, Alignment
    cs = wb.create_sheet(title="_Connections")
    cs.sheet_state = "hidden"

    cs["A1"] = "AtScale MDX Connection"
    cs["A1"].font = Font(bold=True, size=13)

    rows = [
        ("Connection Name",   connection_name),
        ("XMLA Endpoint",     xmla_url),
        ("Initial Catalog",   catalog),
        ("User",              username),
        ("Connection String", conn_string),
        ("", ""),
        ("How to connect",
         "Excel ▶ Data ▶ Get Data ▶ From Other Sources ▶ From Analysis Services"),
        ("",  f"  Server: {xmla_url}"),
        ("",  f"  Database (Catalog): {catalog}"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        cs.cell(row=i, column=1, value=k).font      = Font(bold=True, size=9)
        cs.cell(row=i, column=2, value=v).font      = Font(size=9)
    cs.column_dimensions["A"].width = 24
    cs.column_dimensions["B"].width = 90


# ---------------------------------------------------------------------------
# OLAP XML injection
# ---------------------------------------------------------------------------

def _strip_previous_olap(existing):
    """
    Remove all pivot-table / pivot-cache / connection parts that may have been
    injected by a previous run.  This prevents stale relationship IDs and
    duplicate entries that cause Excel to report the workbook as corrupt when
    the script is run more than once on the same target file.
    """
    pivot_prefixes = (
        "xl/pivotTables/",
        "xl/pivotCache/",
        "xl/connections.xml",
    )
    stale = [k for k in list(existing) if any(k.startswith(p) or k == p
                                               for p in pivot_prefixes)]
    for k in stale:
        del existing[k]

    # workbook.xml.rels — remove pivotCacheDefinition and connections relationships
    wb_rels_key = "xl/_rels/workbook.xml.rels"
    if wb_rels_key in existing:
        src = existing[wb_rels_key].decode("utf-8")
        src = re.sub(
            r'<Relationship\b[^>]*\b(?:pivotCacheDefinition|connections)[^>]*/>', "", src
        )
        existing[wb_rels_key] = src.encode("utf-8")

    # workbook.xml — remove <pivotCaches> section
    wb_key = "xl/workbook.xml"
    if wb_key in existing:
        src = existing[wb_key].decode("utf-8")
        src = re.sub(r'<pivotCaches>.*?</pivotCaches>', "", src)
        existing[wb_key] = src.encode("utf-8")

    # [Content_Types].xml — remove Override entries for pivot / connection parts
    ct_key = "[Content_Types].xml"
    if ct_key in existing:
        src = existing[ct_key].decode("utf-8")
        src = re.sub(
            r'<Override\b[^>]*\b(?:pivotTable|pivotCache|connections)[^>]*/>', "", src
        )
        existing[ct_key] = src.encode("utf-8")

    # Sheet _rels — remove pivotTable relationships from every sheet
    for key in list(existing):
        if re.match(r"xl/worksheets/_rels/sheet\d+\.xml\.rels$", key):
            src = existing[key].decode("utf-8")
            src = re.sub(r'<Relationship\b[^>]*/relationships/pivotTable[^>]*/>', "", src)
            existing[key] = src.encode("utf-8")


def _extract_model_hierarchies(models, model_name):
    """
    Extract OLAP hierarchy info from the model YAML for use in cacheHierarchies
    and pivotHierarchies.  Returns a list of dicts:
      {unique_name, caption, is_measure, display_folder, measure_group,
       default_member_unique_name, all_unique_name, dimension_unique_name}

    Dimensions come from mdx.columns where role == "dimension".
    Measures come from mdx.metrics.

    Falls back to an empty list if the model is not found.
    """
    # Try several name strategies: exact, last component, first component
    model = (
        models.get(model_name) or
        models.get(model_name.split(".")[-1]) or
        models.get(model_name.split(".")[0])
    )
    if not model or not isinstance(model, dict):
        return []

    mdx = model.get("mdx", {})
    if not isinstance(mdx, dict):
        return []

    hierarchies = []
    measure_group = model_name.split(".")[-1]

    # Dimensions — prefer mdx.columns, fall back to sql.columns
    sql        = model.get("sql", {})
    dim_source = mdx.get("columns") or (sql.get("columns") if isinstance(sql, dict) else None)
    if isinstance(dim_source, dict):
        for col_key, col_info in dim_source.items():
            if not isinstance(col_info, dict):
                continue
            if col_info.get("role") != "dimension":
                continue
            label = col_info.get("label", col_key)
            dim_name = label.replace(" ", "_")
            dim_unique = f"[{dim_name}]"
            hier_unique = f"[{dim_name}].[{dim_name} Hierarchy]"
            hierarchies.append({
                "unique_name": hier_unique,
                "caption": f"{dim_name} Hierarchy",
                "default_member_unique_name": f"{hier_unique}.[All]",
                "all_unique_name": f"{hier_unique}.[All]",
                "dimension_unique_name": dim_unique,
                "display_folder": col_info.get("folder", ""),
                "is_measure": False,
                "measure_group": None,
            })

    # Measures — from mdx.metrics
    metrics = mdx.get("metrics", [])
    if isinstance(metrics, list):
        for metric in metrics:
            if not isinstance(metric, dict):
                continue
            query_name = metric.get("query_name", "")
            if not query_name:
                continue
            caption = metric.get("caption", query_name)
            hierarchies.append({
                "unique_name": f"[Measures].[{query_name}]",
                "caption": caption,
                "display_folder": metric.get("folder", ""),
                "is_measure": True,
                "measure_group": measure_group,
                "default_member_unique_name": None,
                "all_unique_name": None,
                "dimension_unique_name": None,
            })

    return hierarchies


def _inject_olap(target_file, pivot_meta, conn_string, catalog, connection_name, models):
    """
    Post-process the xlsx (zip) file to add:
      xl/connections.xml                    – OLEDB/MDX connection definition
      xl/pivotCache/pivotCacheDefinitionN.xml
      xl/pivotTables/pivotTableN.xml
    and update [Content_Types].xml, xl/_rels/workbook.xml.rels, and xl/workbook.xml.

    No pivotCacheRecords files are generated — Excel populates the cache from the
    OLAP cube on the first refresh, matching the structure of correct.xlsx.
    """
    if not pivot_meta:
        return

    tmp = target_file + ".tmp"
    shutil.copy2(target_file, tmp)

    try:
        with zipfile.ZipFile(tmp, "r") as zin, \
             zipfile.ZipFile(target_file, "w", zipfile.ZIP_DEFLATED) as zout:

            existing = {item.filename: zin.read(item.filename)
                        for item in zin.infolist()}

            # Strip any pivot-related content from a previous run so we always
            # start clean.  Stale references in _rels files cause Excel to report
            # the workbook as corrupt.
            _strip_previous_olap(existing)

            new_parts = {}

            # connections.xml — catalog is the MDX cube/command name
            new_parts["xl/connections.xml"] = _connections_xml(
                conn_string, connection_name, catalog
            ).encode("utf-8")

            # pivot cache + table per tile
            for idx, pt in enumerate(pivot_meta, start=1):
                cache_id    = idx - 1  # 0-based cacheId
                model_name  = pt.get("model", "")
                hierarchies = _extract_model_hierarchies(models, model_name)
                measures    = pt.get("measures", [])

                # Resolve the xAxis field name to the matching cacheHierarchy
                # unique name so the pivot table can reference it correctly
                model_obj    = (models.get(model_name) or
                                models.get(model_name.split(".")[-1]) or
                                models.get(model_name.split(".")[0]))
                xaxis_unique = _resolve_xaxis_unique(
                    pt.get("x_axis"), model_obj, hierarchies
                )

                new_parts[f"xl/pivotCache/pivotCacheDefinition{idx}.xml"] = (
                    _cache_def_xml(hierarchies, model_name, measures).encode("utf-8")
                )
                new_parts[f"xl/pivotTables/pivotTable{idx}.xml"] = (
                    _pivot_table_xml(pt, cache_id, hierarchies, xaxis_unique).encode("utf-8")
                )

            # Update workbook.xml.rels — track which rIds were assigned to caches
            wb_rels_key = "xl/_rels/workbook.xml.rels"
            wb_rels_src = existing.get(wb_rels_key, b"").decode("utf-8")
            wb_rels_src, cache_rids = _update_workbook_rels(wb_rels_src, pivot_meta)
            existing[wb_rels_key] = wb_rels_src.encode("utf-8")

            # Update workbook.xml — add <pivotCaches> so Excel discovers the caches
            existing["xl/workbook.xml"] = _update_workbook_xml(
                existing.get("xl/workbook.xml", b""), cache_rids
            )

            # Update Content_Types
            existing["[Content_Types].xml"] = _update_content_types(
                existing.get("[Content_Types].xml", b""), new_parts
            )

            # Update each dashboard sheet _rels to reference pivot tables
            _attach_pivot_tables_to_sheets(existing, pivot_meta)

            existing.update(new_parts)
            for fname, data in existing.items():
                zout.writestr(fname, data)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


def _connections_xml(conn_string, connection_name, catalog):
    # type="5"  = OLE DB (MSOLAP is an OLE DB provider)
    # commandType="1" = cube/MDX; command = cube name (catalog)
    # olapPr is required by Excel for OLAP connections
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<connection id="1" name="{_xml_attr(connection_name)}" '
        'type="5" refreshedVersion="8" background="1">'
        f'<dbPr connection="{_xml_attr(conn_string)}" '
        f'command="{_xml_attr(catalog)}" commandType="1"/>'
        '<olapPr sendLocale="1" rowDrillCount="1000"/>'
        '</connection>'
        '</connections>'
    )


def _resolve_xaxis_unique(x_axis, model, hierarchies):
    """
    Convert a namespace xAxis field name (e.g. "query_hour") to its
    cacheHierarchy unique name (e.g. "[Query_Hour].[Query_Hour Hierarchy]").
    Returns None if the field cannot be resolved.
    """
    if not x_axis:
        return None

    # Strategy 1: look up the column key in model sql.columns, use its label
    if model and isinstance(model, dict):
        sql  = model.get("sql", {})
        cols = sql.get("columns", {}) if isinstance(sql, dict) else {}
        col  = cols.get(x_axis) if isinstance(cols, dict) else None
        if col and isinstance(col, dict):
            label     = col.get("label", x_axis)
            dim_name  = label.replace(" ", "_")
            candidate = f"[{dim_name}].[{dim_name} Hierarchy]"
            if any(h["unique_name"] == candidate for h in hierarchies):
                return candidate

    # Strategy 2: fuzzy match — normalise underscores/spaces
    x_norm = x_axis.lower().replace("_", "").replace(" ", "")
    for h in hierarchies:
        if not h["is_measure"]:
            dim_norm = h["dimension_unique_name"].strip("[]").lower().replace("_", "").replace(" ", "")
            if x_norm == dim_norm:
                return h["unique_name"]

    return None


def _cache_def_xml(hierarchies, model_name, measures):
    """
    Emit an OLAP pivot cache definition matching the structure Excel generates
    after connecting to an OLAP cube (as seen in correct.xlsx).

    cacheFields contains one entry per selected measure with the hierarchy index
    so Excel can pre-populate the field list without a full refresh.
    cacheHierarchies lists all known OLAP hierarchies from the model YAML.
    """
    r  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    n_hier = len(hierarchies)
    dims   = [h for h in hierarchies if not h["is_measure"]]
    meas   = [h for h in hierarchies if h["is_measure"]]

    # cacheFields: one per selected measure (dimensions are referenced via
    # cacheHierarchies, not cacheFields, in OLAP mode)
    cf_parts = []
    for m_name in measures:
        m_unique = f"[Measures].[{m_name}]"
        m_idx    = next((i for i, h in enumerate(hierarchies)
                         if h["unique_name"] == m_unique), -1)
        if m_idx < 0:
            continue
        m_caption = next((h["caption"] for h in hierarchies
                          if h["unique_name"] == m_unique), m_name)
        cf_parts.append(
            f'<cacheField name="{_xml_attr(m_unique)}" '
            f'caption="{_xml_attr(m_caption)}" '
            f'numFmtId="0" hierarchy="{m_idx}" level="32767"/>'
        )

    n_cf = len(cf_parts)
    cache_fields_xml = (
        f'<cacheFields count="{n_cf}">{"".join(cf_parts)}</cacheFields>'
        if n_cf > 0 else '<cacheFields count="0"/>'
    )

    # cacheHierarchies XML
    hier_parts = []
    for h in hierarchies:
        if h["is_measure"]:
            hier_parts.append(
                f'<cacheHierarchy uniqueName="{_xml_attr(h["unique_name"])}" '
                f'caption="{_xml_attr(h["caption"])}" '
                f'measure="1" displayFolder="{_xml_attr(h["display_folder"])}" '
                f'measureGroup="{_xml_attr(h["measure_group"])}" count="0"/>'
            )
        else:
            hier_parts.append(
                f'<cacheHierarchy uniqueName="{_xml_attr(h["unique_name"])}" '
                f'caption="{_xml_attr(h["caption"])}" '
                f'defaultMemberUniqueName="{_xml_attr(h["default_member_unique_name"])}" '
                f'allUniqueName="{_xml_attr(h["all_unique_name"])}" '
                f'dimensionUniqueName="{_xml_attr(h["dimension_unique_name"])}" '
                f'displayFolder="{_xml_attr(h["display_folder"])}" '
                'count="0" unbalanced="0"/>'
            )

    # dimensions XML: Measures dimension first, then each non-measure dimension
    group_name = _xml_attr(model_name.split(".")[-1])
    dim_parts  = ['<dimension measure="1" name="Measures" uniqueName="[Measures]" caption="Measures"/>']
    for h in dims:
        dim_name = _xml_attr(h["dimension_unique_name"].strip("[]"))
        dim_parts.append(
            f'<dimension name="{dim_name}" '
            f'uniqueName="{_xml_attr(h["dimension_unique_name"])}" '
            f'caption="{dim_name}"/>'
        )

    map_parts = [f'<map measureGroup="0" dimension="{i + 1}"/>' for i in range(len(dims))]
    n_dims    = 1 + len(dims)

    cache_hier_xml = (
        f'<cacheHierarchies count="{n_hier}">{"".join(hier_parts)}</cacheHierarchies>'
        if n_hier > 0 else '<cacheHierarchies count="0"/>'
    )
    dims_xml = (
        f'<dimensions count="{n_dims}">{"".join(dim_parts)}</dimensions>'
        if n_dims > 0 else ''
    )
    mg_xml = (
        f'<measureGroups count="1">'
        f'<measureGroup name="{group_name}" caption="{group_name}"/>'
        f'</measureGroups>'
        if meas else ''
    )
    maps_xml = (
        f'<maps count="{len(dims)}">{"".join(map_parts)}</maps>'
        if dims else ''
    )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<pivotCacheDefinition xmlns="{ns}" xmlns:r="{r}" '
        'saveData="0" backgroundQuery="1" '
        'createdVersion="8" refreshedVersion="8" minRefreshableVersion="3" '
        'recordCount="0" supportSubquery="1" supportAdvancedDrill="1">'
        '<cacheSource type="external" connectionId="1"/>'
        + cache_fields_xml +
        cache_hier_xml +
        '<kpis count="0"/>'
        + dims_xml + mg_xml + maps_xml +
        '</pivotCacheDefinition>'
    )


def _pivot_table_xml(pt, cache_id, hierarchies, xaxis_unique=None):
    """
    Emit an OLAP pivot table definition pre-configured with the namespace
    measures and xAxis, matching the structure Excel generates (correct.xlsx).

    Element order per OOXML schema:
      location → pivotFields → rowFields → rowItems → colItems →
      dataFields → pivotHierarchies → pivotTableStyleInfo
    """
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    tile_title = _xml_attr(pt["tile_title"])
    anchor_col = pt["anchor_col"]
    hdr_row    = pt["hdr_row"]
    measures   = pt.get("measures", [])
    has_xaxis  = bool(xaxis_unique)
    n_hier     = len(hierarchies)

    # ---- location ----
    if has_xaxis:
        # Single-cell anchor; firstHeaderRow/DataRow=0 — pivot table writes its
        # own header at hdr_row and data below (matches user's corrected workbook)
        loc_xml = (
            f'<location ref="{_col(anchor_col)}{hdr_row}" '
            'firstHeaderRow="0" firstDataRow="0" firstDataCol="0"/>'
        )
    else:
        # Text / total-only: two rows — header + single value
        loc_xml = (
            f'<location ref="{_col(anchor_col)}{hdr_row}:{_col(anchor_col)}{hdr_row + 1}" '
            'firstHeaderRow="1" firstDataRow="1" firstDataCol="0"/>'
        )

    # ---- pivotFields ----
    # One entry per measure (index 0..N-1), then one for the row dimension
    pf_parts = []
    for _ in measures:
        pf_parts.append('<pivotField dataField="1" showAll="0"/>')
    if has_xaxis:
        pf_parts.append(
            '<pivotField axis="axisRow" allDrilled="1" showAll="0" dataSourceSort="1"/>'
        )
    pivot_fields_xml = (
        f'<pivotFields count="{len(pf_parts)}">{"".join(pf_parts)}</pivotFields>'
        if pf_parts else ''
    )

    # ---- rowFields (only when xAxis present) ----
    # The dimension pivotField is at index len(measures) in pivotFields
    row_fields_xml = (
        f'<rowFields count="1"><field x="{len(measures)}"/></rowFields>'
        if has_xaxis else ''
    )

    # ---- rowItems / colItems ----
    # Always required when dataFields is present.
    # For xAxis tiles: grand total row only (data rows populated on refresh).
    # For text tiles: single row (the total).
    row_items_xml = col_items_xml = ''
    if measures:
        if has_xaxis:
            # One grand-total row; actual data rows appear after refresh
            row_items_xml = '<rowItems count="1"><i t="grand"><x/></i></rowItems>'
        else:
            row_items_xml = '<rowItems count="1"><i/></rowItems>'
        col_items_xml = '<colItems count="1"><i/></colItems>'

    # ---- dataFields ----
    data_fields_xml = (
        '<dataFields count="1"><dataField fld="0" baseField="0" baseItem="0"/></dataFields>'
        if measures else ''
    )

    # ---- pivotHierarchies ----
    ph_parts = []
    for h in hierarchies:
        if h["is_measure"]:
            ph_parts.append(
                '<pivotHierarchy dragToRow="0" dragToCol="0" dragToPage="0" dragToData="1"/>'
            )
        else:
            ph_parts.append('<pivotHierarchy/>')
    pivot_hier_xml = (
        f'<pivotHierarchies count="{n_hier}">{"".join(ph_parts)}</pivotHierarchies>'
        if n_hier > 0 else ''
    )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<pivotTableDefinition xmlns="{ns}" '
        f'name="{tile_title}" cacheId="{cache_id}" '
        'applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" '
        'applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" '
        'dataCaption="Values" updatedVersion="8" minRefreshableVersion="3" '
        'useAutoFormatting="1" itemPrintTitles="1" createdVersion="8" '
        'indent="0" outline="1" outlineData="1" '
        'multipleFieldFilters="0" fieldListSortAscending="1">'
        + loc_xml
        + pivot_fields_xml
        + row_fields_xml
        + row_items_xml
        + col_items_xml
        + data_fields_xml
        + pivot_hier_xml
        + '<pivotTableStyleInfo name="PivotStyleMedium9" '
        'showRowHeaders="1" showColHeaders="1" showRowStripes="0" '
        'showColStripes="0" showLastColumn="1"/>'
        '</pivotTableDefinition>'
    )


def _update_content_types(raw, new_parts):
    """Append <Override> entries for each new part."""
    if not raw:
        return raw
    try:
        src = raw.decode("utf-8")
        type_map = {
            "pivotCacheDefinition": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.pivotCacheDefinition+xml"
            ),
            "pivotCacheRecords": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.pivotCacheRecords+xml"
            ),
            "pivotTable": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.pivotTable+xml"
            ),
            "connections": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.connections+xml"
            ),
        }
        inserts = []
        for part_path in new_parts:
            part_name = "/" + part_path
            if part_name in src:
                continue
            ct = None
            for key, mime in type_map.items():
                if key in part_path:
                    ct = mime
                    break
            if ct:
                inserts.append(
                    f'<Override PartName="{part_name}" ContentType="{ct}"/>'
                )
        if inserts:
            src = src.replace("</Types>", "".join(inserts) + "</Types>")
        return src.encode("utf-8")
    except Exception as e:
        print(f"Warning: content-types update failed: {e}", file=sys.stderr)
        return raw


def _update_workbook_rels(src, pivot_meta):
    """
    Add relationships for pivot caches and the connections file.
    Returns (updated_src_str, cache_rids) where cache_rids is {idx: rId_str}.
    """
    cache_rids = {}
    try:
        next_id = _next_rid(src)
        inserts = []
        cache_type = (
            "http://schemas.openxmlformats.org/officeDocument/2006"
            "/relationships/pivotCacheDefinition"
        )
        conn_type = (
            "http://schemas.openxmlformats.org/officeDocument/2006"
            "/relationships/connections"
        )

        for idx in range(1, len(pivot_meta) + 1):
            target = f"pivotCache/pivotCacheDefinition{idx}.xml"
            if target not in src:
                rid = f"rId{next_id}"
                cache_rids[idx] = rid
                inserts.append(
                    f'<Relationship Id="{rid}" Type="{cache_type}" '
                    f'Target="{target}"/>'
                )
                next_id += 1

        if "connections.xml" not in src:
            inserts.append(
                f'<Relationship Id="rId{next_id}" Type="{conn_type}" '
                'Target="connections.xml"/>'
            )

        if inserts:
            src = src.replace("</Relationships>", "".join(inserts) + "</Relationships>")
        return src, cache_rids
    except Exception as e:
        print(f"Warning: workbook rels update failed: {e}", file=sys.stderr)
        return src, cache_rids


def _update_workbook_xml(raw, cache_rids):
    """
    Add a <pivotCaches> section to workbook.xml so Excel can discover and
    register the pivot caches.  Each <pivotCache cacheId="N" r:id="rIdXX"/>
    entry uses the relationship ID that was added to workbook.xml.rels.
    """
    if not raw or not cache_rids:
        return raw
    try:
        src = raw.decode("utf-8")
        if "<pivotCaches" in src:
            return raw  # already present

        r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        cache_entries = "".join(
            f'<pivotCache cacheId="{idx - 1}" r:id="{rid}"/>'
            for idx, rid in sorted(cache_rids.items())
        )
        pivot_caches_xml = f'<pivotCaches>{cache_entries}</pivotCaches>'

        # openpyxl's workbook.xml should already have xmlns:r, but be safe
        if 'xmlns:r=' not in src:
            src = src.replace('<workbook ', f'<workbook xmlns:r="{r_ns}" ', 1)

        src = src.replace("</workbook>", pivot_caches_xml + "</workbook>")
        return src.encode("utf-8")
    except Exception as e:
        print(f"Warning: workbook.xml update failed: {e}", file=sys.stderr)
        return raw


def _attach_pivot_tables_to_sheets(existing, pivot_meta):
    """
    For each pivot table:
      - Add a Relationship in the sheet's _rels file pointing to the pivot table
      - Create pivotTable _rels (pivotTable → cacheDefinition)

    No pivotCache _rels are created (no pivotCacheRecords file is used —
    Excel fetches cache data from the OLAP cube on refresh, matching correct.xlsx).

    NOTE: pivot tables must NOT be listed in <tableParts> — that element is for
    regular Excel Tables (ListObjects).  The sheet-to-pivotTable link is made
    exclusively through the sheet's _rels file.
    """
    pkg_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    pt_rel_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006"
        "/relationships/pivotTable"
    )
    pt_cache_rel_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006"
        "/relationships/pivotCacheDefinition"
    )

    # Build a map of sheet title -> sheet file key
    sheet_map = _build_sheet_map(existing)

    for idx, pt in enumerate(pivot_meta, start=1):
        sheet_title = pt["sheet_title"]
        sheet_key   = sheet_map.get(sheet_title)
        if not sheet_key:
            continue

        # --- sheet _rels: add relationship from sheet to pivot table ---
        sheet_num = re.search(r"sheet(\d+)\.xml$", sheet_key)
        sheet_num = sheet_num.group(1) if sheet_num else "1"
        rels_key  = f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels"

        rels_src = existing.get(rels_key, b"").decode("utf-8")
        if not rels_src.strip():
            rels_src = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                f'<Relationships xmlns="{pkg_ns}"></Relationships>'
            )
        # Convert self-closing <Relationships .../> to open/close form so we can inject
        rels_src = re.sub(
            r'(<Relationships\b[^>]*)/>', r'\1></Relationships>', rels_src
        )

        if f"pivotTable{idx}.xml" not in rels_src:
            next_id   = _next_rid(rels_src)
            rel_entry = (
                f'<Relationship Id="rId{next_id}" Type="{pt_rel_type}" '
                f'Target="../pivotTables/pivotTable{idx}.xml"/>'
            )
            rels_src = rels_src.replace(
                "</Relationships>", rel_entry + "</Relationships>"
            )
        existing[rels_key] = rels_src.encode("utf-8")

        # NOTE: No pivotCacheRecords file or pivotCache _rels are generated.
        # Excel populates the cache from the OLAP cube on the first refresh
        # (matching the structure of correct.xlsx which has no _rels for caches).

        # --- pivotTable _rels: pivotTable → cacheDefinition ---
        pt_rels_key = f"xl/pivotTables/_rels/pivotTable{idx}.xml.rels"
        existing[pt_rels_key] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{pkg_ns}">'
            f'<Relationship Id="rId1" Type="{pt_cache_rel_type}" '
            f'Target="../pivotCache/pivotCacheDefinition{idx}.xml"/>'
            '</Relationships>'
        ).encode("utf-8")


def _build_sheet_map(existing):
    """Return {sheetTitle: sheetXmlKey} by parsing workbook.xml + workbook.xml.rels."""
    wb_xml   = existing.get("xl/workbook.xml", b"").decode("utf-8")
    rels_xml = existing.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8")
    sheet_map = {}

    for tag_m in re.finditer(r'<sheet\b([^>]+)>', wb_xml):
        attrs = tag_m.group(1)
        name_m = re.search(r'\bname="([^"]+)"', attrs)
        rid_m  = re.search(r'\br:id="([^"]+)"', attrs)
        if not name_m or not rid_m:
            continue
        name, rid = name_m.group(1), rid_m.group(1)

        # Find target in rels — attribute order may vary
        target_m = re.search(
            r'<Relationship\b[^>]*\bId="' + re.escape(rid) + r'"[^>]*\bTarget="([^"]+)"',
            rels_xml,
        )
        if not target_m:
            target_m = re.search(
                r'<Relationship\b[^>]*\bTarget="([^"]+)"[^>]*\bId="' + re.escape(rid) + r'"',
                rels_xml,
            )
        if target_m:
            target = target_m.group(1).lstrip("/")
            # Targets are relative to xl/ (e.g. "worksheets/sheet1.xml")
            sheet_key = target if target.startswith("xl/") else f"xl/{target}"
            sheet_map[name] = sheet_key

    return sheet_map


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_name(name):
    for ch in r'/\?*[]:':
        name = name.replace(ch, "-")
    return name[:31]


def _col(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


def _xml_attr(s):
    return (str(s)
            .replace("&", "&amp;")
            .replace('"', "&quot;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def _next_rid(xml_src):
    ids = [int(m) for m in re.findall(r'Id="rId(\d+)"', xml_src)]
    return max(ids, default=0) + 1


if __name__ == "__main__":
    main()
